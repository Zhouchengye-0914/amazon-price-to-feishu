# -*- coding: utf-8 -*-
"""feishu.py — 飞书 API：token/wiki/读原始表/同步基础数据/六列稀疏写入/一次性列迁移

流程（方案 11）：先同步原始周报 → 同一快照抓取 → 六列写回目标表。
写入安全（方案 14）：按 Sheet+ASIN 定位行，只写有结果的行，连续行合并小 Range。
"""
from __future__ import annotations

import re
import time
import hashlib
import json
from decimal import Decimal, InvalidOperation
from pathlib import Path

import httpx

from config import BASE_DIR
from models import (
    CrawlResult, ReportRow, TARGET_SOURCE_EXCEL, TARGET_SOURCE_FALLBACK,
    TARGET_SOURCE_FEISHU, TARGET_SOURCE_MISSING,
)
from pricing import calc_target_price, dec
from weekly_registry import parse_registry_values

FEISHU_BASE = 'https://open.feishu.cn/open-apis'
# 旧版追加列表头（迁移时清理）
LEGACY_HEADERS = ['抓取的价格', '抓取的价格折扣', '一致性检查', '抓取时间戳']
COMPACT_BASE_HEADERS = ['ASIN', 'SKU', '尺寸', '正常售价', '本周折扣形式', '本周折扣%', '目标成交价']

# 列名 → 字段键（表头自动定位用）
_COL_NAME_MAP = {
    'asin': 'ASIN',
    'sku': 'SKU',
    'size': '尺寸',
    'normal_price': '正常售价',
    'h_type': '本周折扣形式',
    'i_value': '本周折扣%',
    'target_price': '目标成交价',
}


def load_migration_layout() -> dict[str, int]:
    """读迁移记录中的 output_layout（{sheet: start_col}）；无记录返回空"""
    from config import OUTPUT_DIR
    p = OUTPUT_DIR / 'migration_record.json'
    if not p.exists():
        return {}
    try:
        import json
        with open(p, encoding='utf-8') as f:
            return json.load(f).get('output_layout') or {}
    except Exception:
        return {}


def save_migration_layout(layout: dict[str, int]) -> None:
    """把解析出的输出布局写入迁移记录（日常运行优先读它）"""
    from config import OUTPUT_DIR
    p = OUTPUT_DIR / 'migration_record.json'
    try:
        import json
        data = json.loads(p.read_text(encoding='utf-8')) if p.exists() else {}
        data['output_layout'] = layout
        with open(p, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=1)
    except Exception:
        pass


def _find_existing_six(header_row: list, headers: list[str]) -> int | None:
    """表头里已存在完整六列（顺序一致）→ 返回起始列(1-based)，否则 None"""
    n = len(headers)
    for i, cell in enumerate(header_row):
        if not (isinstance(cell, str) and cell.strip() == headers[0]):
            continue
        if all(isinstance(header_row[i + j], str) and header_row[i + j].strip() == headers[j]
               for j in range(n) if i + j < len(header_row)):
            return i + 1
    return None


def resolve_sheet_start_col(header_row: list, sample_rows: list[list],
                            headers: list[str]) -> int | None:
    """决定某 sheet 的六列起始列（1-based）：
    1) 表头已存在完整六列 → 复用其起始列；
    2) 否则找最后业务列之后连续 6 个空列（表头+数据都空）。
    返回 None 表示找不到。"""
    # 1. 已有六列表头
    existing = _find_existing_six(header_row, headers)
    if existing:
        return existing
    # 2. 最后业务列
    last_biz = 0
    for i, cell in enumerate(header_row):
        if isinstance(cell, str) and cell.strip():
            last_biz = i + 1
    # 3. 从最后业务列之后找连续 6 空列
    col = last_biz + 1
    max_col = 400
    while col + len(headers) - 1 <= max_col:
        ok = True
        for j in range(len(headers)):
            c = col + j
            if _col_has_content(header_row, sample_rows, c):
                ok = False
                break
        if ok:
            return col
        col += 1
    return None


def _col_has_content(header_row: list, sample_rows: list[list], col: int) -> bool:
    """列是否有内容：表头行或任何样例行非空"""
    for row in [header_row] + list(sample_rows[:8]):
        if col - 1 < len(row) and row[col - 1] is not None:
            if isinstance(row[col - 1], str):
                if row[col - 1].strip() != '':
                    return True
            else:
                return True
    return False


def col_letter(n: int) -> str:
    s = ''
    while n > 0:
        n, rem = divmod(n - 1, 26)
        s = chr(65 + rem) + s
    return s


def _detect_header_row(vals: list[list]) -> int:
    """在前几行里找含 'ASIN' 的表头行（0-based），找不到返回 -1"""
    from weekly_mapping import cell_text
    for i, row in enumerate(vals[:10]):
        for cell in row:
            if cell_text(cell).upper() == 'ASIN':
                return i
    return -1


def _resolve_cols(header_row: list, cfg: dict) -> dict:
    """按表头名定位列（1-based），找不到的字段用 config.source_cols 兜底"""
    cols = dict(cfg['source_cols'])
    from weekly_mapping import cell_text
    for i, cell in enumerate(header_row):
        s = cell_text(cell)
        for key, name in _COL_NAME_MAP.items():
            if s.casefold() == name.casefold():
                cols[key] = i + 1
    return cols


def read_source_rows(vals: list[list], cfg: dict,
                     source_kind: str = 'feishu') -> tuple[list[ReportRow], list[dict]]:
    """从 2D 单元格矩阵解析周报行（表头自动探测 + 列名定位 + target 本地兜底）。
    返回 (有效行, 无效行[sheet 由调用方补])。
    - K 列目标成交价为公式且读不到计算值时，用本地 calc_target_price 兜底；
    - 目标成交价/正常售价仍为空 → 判为无效行（3.7）；
    - target_price_source 追踪：feishu_value / excel_cached_value / local_fallback / missing。"""
    hr = _detect_header_row(vals)
    if hr < 0:
        raise RuntimeError('未找到含 ASIN 的表头行，请检查源表结构')
    cols = _resolve_cols(vals[hr], cfg)
    rows: list[ReportRow] = []
    invalid: list[dict] = []
    for idx, row in enumerate(vals[hr + 1:], start=hr + 2):
        def _cell(c):
            return row[c - 1] if len(row) >= c else None
        asin_v = _cell(cols['asin'])
        asin = str(asin_v).strip() if asin_v is not None else ''
        if not asin.startswith('B0'):
            from product_links import normalize_product, ProductLinkError
            try:
                asin, _ = normalize_product(asin_v, cfg.get('source_marketplace', 'US'))
            except ProductLinkError as exc:
                if 'http' in str(asin_v).lower():
                    raise RuntimeError(f'源行{idx}商品链接无效: {exc}') from exc
                continue
        rr = ReportRow(
            row_num=idx, asin=asin,
            sku=str(_cell(cols['sku']) or '').strip(),
            size=str(_cell(cols.get('size', 4)) or '').strip(),
            normal_price=_num_or_none(_cell(cols['normal_price'])),
            h_type=str(_cell(cols['h_type']) or '').strip(),
            i_value=_parse_i_value(_cell(cols['i_value'])),
        )
        k = _num_or_none(_cell(cols['target_price']))
        if k is not None:
            rr.target_price = k
            rr.target_price_source = (TARGET_SOURCE_FEISHU if source_kind == 'feishu'
                                      else TARGET_SOURCE_EXCEL)
        else:
            # 公式无缓存值 → 本地计算兜底
            rr.target_price = calc_target_price(rr)
            rr.target_price_source = (TARGET_SOURCE_FALLBACK if rr.target_price is not None
                                      else TARGET_SOURCE_MISSING)
        missing = []
        for name, value in (('目标成交价', rr.target_price), ('正常售价', rr.normal_price)):
            if value is not None and (not value.is_finite() or value < 0):
                missing.append(f'{name}必须是有限非负数')
        if rr.target_price is None:
            missing.append('目标成交价(K)')
        if rr.normal_price is None:
            missing.append('正常售价(E)')
        if missing:
            invalid.append({'row_num': idx, 'asin': asin,
                            'reason': '、'.join(missing) + ' 为空',
                            'report_row': rr})
            continue
        rows.append(rr)
    return rows, invalid


def _json_safe(v):
    """JSON 序列化兼容：Decimal → float，list 递归"""
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, list):
        return [_json_safe(x) for x in v]
    return v


def _num_or_none(v):
    """飞书单元格 → Decimal/文本（公式结果按数值处理）"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        try:
            return Decimal(str(v))
        except InvalidOperation:
            return None
    s = str(v).strip().replace('￥', '').replace('$', '').replace(',', '').replace('%', '')
    if s == '':
        return None
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


class FeishuClient:
    def __init__(self, cfg: dict):
        self.cfg = cfg
        self.token = ''
        self._token_expires_at = 0.0
        self._client = httpx.Client(
            base_url=FEISHU_BASE,
            timeout=httpx.Timeout(20, connect=15),
            trust_env=False,
        )

    # ---------- 认证与解析 ----------
    def auth(self) -> str:
        # Refresh proactively. A client lives for more than an hour during a full
        # run; caching a tenant token forever makes the final notification fail.
        expires_at = getattr(self, '_token_expires_at', 0.0)
        if self.token and (expires_at <= 0 or time.monotonic() < expires_at):
            return self.token
        r = self._client.post('/auth/v3/tenant_access_token/internal', json={
            'app_id': self.cfg['feishu_app_id'],
            'app_secret': self.cfg['feishu_app_secret'],
        })
        r.raise_for_status()
        d = r.json()
        if d.get('code') != 0:
            raise RuntimeError(f"飞书认证失败: {d.get('msg')}")
        self.token = d['tenant_access_token']
        expires_in = max(0, int(d.get('expire') or d.get('expires_in') or 7200))
        self._token_expires_at = time.monotonic() + max(0, expires_in - 300)
        return self.token

    def _headers(self) -> dict:
        return {'Authorization': f'Bearer {self.auth()}'}

    @staticmethod
    def resolve_wiki_url(wiki_url: str) -> str:
        m = re.search(r'/wiki/([A-Za-z0-9_-]+)', wiki_url)
        if not m:
            raise RuntimeError(f'wiki 链接格式错误: {wiki_url}')
        return m.group(1)

    def resolve_wiki_obj(self, wiki_url: str) -> tuple[str, str]:
        """wiki token → (obj_token, obj_type)。obj_type: sheet/bitable/file 等"""
        node_token = self.resolve_wiki_url(wiki_url)
        r = self._client.get('/wiki/v2/spaces/get_node',
                             params={'token': node_token}, headers=self._headers())
        r.raise_for_status()
        d = r.json()
        if d.get('code') != 0:
            raise RuntimeError(f"wiki 解析失败: {d.get('msg')}")
        node = (d.get('data') or {}).get('node') or {}
        obj = node.get('obj_token') or ''
        typ = node.get('obj_type') or 'sheet'
        if not obj:
            raise RuntimeError(f'wiki 节点没有对象: {wiki_url}')
        return obj, typ

    def resolve_wiki(self, wiki_url: str) -> str:
        """wiki token → spreadsheet/file token"""
        obj, _ = self.resolve_wiki_obj(wiki_url)
        return obj

    def ensure_permission_member(self, file_token: str, file_type: str,
                                 member_id: str, member_type: str = 'openid',
                                 perm: str = 'full_access') -> dict:
        """幂等确保指定成员拥有云文档权限。"""
        if not file_token or not member_id:
            raise RuntimeError('云文档授权缺少 file_token 或 member_id')
        r = self._client.get(
            f'/drive/v1/permissions/{file_token}/members',
            params={'type': file_type, 'page_size': 100}, headers=self._headers())
        r.raise_for_status()
        data = r.json()
        if data.get('code') != 0:
            raise RuntimeError(f"读取云文档协作者失败: {data.get('msg')}")
        for item in (data.get('data') or {}).get('items') or []:
            if (item.get('member_id') == member_id
                    and item.get('member_type') == member_type
                    and item.get('perm') == perm):
                return {'member_id': member_id, 'member_type': member_type,
                        'perm': perm, 'reused': True}
        r = self._client.post(
            f'/drive/v1/permissions/{file_token}/members',
            params={'type': file_type, 'need_notification': 'true'},
            json={'member_type': member_type, 'member_id': member_id, 'perm': perm},
            headers=self._headers())
        r.raise_for_status()
        data = r.json()
        if data.get('code') != 0:
            raise RuntimeError(f"添加云文档管理协作者失败: {data.get('msg')}")
        return {'member_id': member_id, 'member_type': member_type,
                'perm': perm, 'reused': False}

    def send_text_message(self, open_id: str, text: str) -> str:
        """向指定用户发送任务完成通知，并返回 message_id。"""
        if not open_id or not text.strip():
            raise RuntimeError('飞书通知缺少 open_id 或消息正文')
        r = self._client.post(
            '/im/v1/messages', params={'receive_id_type': 'open_id'},
            json={'receive_id': open_id, 'msg_type': 'text',
                  'content': json.dumps({'text': text}, ensure_ascii=False)},
            headers=self._headers())
        data = r.json()
        if r.is_error:
            raise RuntimeError(f"飞书消息失败: HTTP {r.status_code}, code={data.get('code')}, {data.get('msg')}")
        if data.get('code') != 0:
            raise RuntimeError(f"飞书完成通知失败: {data.get('msg')}")
        message_id = ((data.get('data') or {}).get('message_id') or '')
        if not message_id:
            raise RuntimeError('飞书完成通知未返回 message_id')
        return message_id

    def send_post_message(self, open_id: str, post: dict) -> str:
        """Send a rich-text completion with explicit named links."""
        if not open_id or not post:
            raise RuntimeError('飞书通知缺少 open_id 或消息正文')
        r = self._client.post(
            '/im/v1/messages', params={'receive_id_type': 'open_id'},
            json={'receive_id': open_id, 'msg_type': 'post',
                  'content': json.dumps(post, ensure_ascii=False)},
            headers=self._headers())
        data = r.json()
        if r.is_error or data.get('code') != 0:
            raise RuntimeError(f"飞书消息失败: HTTP {r.status_code}, code={data.get('code')}, {data.get('msg')}")
        message_id = (data.get('data') or {}).get('message_id')
        if not message_id:
            raise RuntimeError('飞书完成通知未返回 message_id')
        return message_id

    def application_collaborators(self) -> list[str]:
        """Application developer collaborators, not document sharing members."""
        r = self._client.get(
            f'/application/v6/applications/{self.cfg["feishu_app_id"]}/collaborators',
            params={'user_id_type': 'open_id'}, headers=self._headers())
        r.raise_for_status()
        data = r.json()
        if data.get('code') != 0:
            raise RuntimeError(f"读取应用协作者失败: {data.get('msg')}")
        ids = [item['user_id'] for item in (data.get('data') or {}).get('collaborators', [])
               if item.get('user_id')]
        if not ids or any(not item.startswith('ou_') for item in ids):
            raise RuntimeError('应用协作者列表为空或不是Open ID，禁止误发')
        return list(dict.fromkeys(ids))

    def rename_spreadsheet(self, token: str, title: str) -> None:
        r = self._client.patch(f'/sheets/v3/spreadsheets/{token}',
                               json={'title': title}, headers=self._headers())
        r.raise_for_status()
        data = r.json()
        if data.get('code') != 0:
            raise RuntimeError(f"更新结果表名称失败: {data.get('msg')}")

    # ---------- 上传文件（原始表为 xlsx file 时） ----------
    def read_source_file(self, file_token: str, sheets: list[str],
                         cfg: dict) -> tuple[dict[str, list[ReportRow]], dict, list[dict]]:
        """原始表是上传的 xlsx 文件（obj_type=file）→ 下载 → openpyxl 解析。
        公式列读缓存值（data_only），缺失走本地 calc_target_price 兜底。"""
        import json as _json
        import tempfile
        import openpyxl
        r = self._client.get(f'/drive/v1/files/{file_token}/download',
                             headers=self._headers())
        r.raise_for_status()
        content = r.content
        if content[:1] == b'{':
            try:
                err = _json.loads(content)
                if err.get('code') != 0:
                    raise RuntimeError(f"下载原始表失败: {err.get('msg')}")
            except RuntimeError:
                raise
            except Exception:
                pass
        tmp = Path(tempfile.mkstemp(suffix='.xlsx')[1])
        try:
            tmp.write_bytes(content)
            wb = openpyxl.load_workbook(tmp, data_only=True)
        except Exception as e:
            raise RuntimeError(f'原始表 xlsx 解析失败: {e}')
        out: dict[str, list[ReportRow]] = {}
        invalid_all: list[dict] = []
        meta = {'source_file': file_token, 'source_type': 'file',
                'read_at': time.strftime('%Y-%m-%d %H:%M:%S')}
        for sheet in sheets:
            if sheet not in wb.sheetnames:
                continue
            ws = wb[sheet]
            vals = [[ws.cell(r, c).value for c in range(1, 16)]
                    for r in range(1, ws.max_row + 1)]
            try:
                rows, invalid = read_source_rows(vals, cfg, source_kind='excel')
            except RuntimeError as e:
                raise RuntimeError(f'[{sheet}] {e}')
            for iv in invalid:
                iv['sheet'] = sheet
            invalid_all.extend(invalid)
            if rows:
                out[sheet] = rows
                meta['sheets'] = meta.get('sheets', {})
                meta['sheets'][sheet] = {'rows': len(rows), 'invalid': len(invalid)}
        return out, meta, invalid_all

    # ---------- 表格 ----------
    def list_sheets(self, spreadsheet: str) -> dict[str, str]:
        r = self._client.get(f'/sheets/v3/spreadsheets/{spreadsheet}/sheets/query',
                             headers=self._headers())
        r.raise_for_status()
        d = r.json()
        if d.get('code') != 0:
            raise RuntimeError(f"sheet 列表失败: {d.get('msg')}")
        return {s.get('title'): s.get('sheet_id')
                for s in ((d.get('data') or {}).get('sheets') or []) if s.get('sheet_id')}

    def query_sheets(self, spreadsheet: str) -> list[dict]:
        """读取完整子表元数据；大表 v3 超时/5xx 时回退 v2 metainfo。"""
        try:
            r = self._client.get(
                f'/sheets/v3/spreadsheets/{spreadsheet}/sheets/query',
                headers=self._headers(), timeout=8)
            if r.status_code < 500:
                r.raise_for_status()
                d = r.json()
                if d.get('code') != 0:
                    raise RuntimeError(f"sheet 列表失败: {d.get('msg')}")
                sheets = ((d.get('data') or {}).get('sheets') or [])
                if any(not (s.get('grid_properties') or {}).get('row_count') for s in sheets):
                    raise RuntimeError('子表元数据缺少行容量，禁止按固定2000行截断')
                return sheets
        except httpx.TransportError:
            pass

        r = self._client.get(f'/sheets/v2/spreadsheets/{spreadsheet}/metainfo',
                             headers=self._headers(), timeout=30)
        r.raise_for_status()
        d = r.json()
        if d.get('code') != 0:
            raise RuntimeError(f"sheet 元数据读取失败: {d.get('msg')}")
        normalized = []
        for sheet in ((d.get('data') or {}).get('sheets') or []):
            normalized.append({
                'sheet_id': sheet.get('sheetId') or sheet.get('sheet_id'),
                'title': sheet.get('title') or '',
                'index': sheet.get('index'),
                'grid_properties': {
                    'row_count': sheet.get('rowCount') or sheet.get('row_count'),
                    'column_count': sheet.get('columnCount') or sheet.get('column_count'),
                },
            })
        if any(not s['grid_properties'].get('row_count') for s in normalized):
            raise RuntimeError('子表元数据缺少行容量，禁止按固定2000行截断')
        return normalized

    def spreadsheet_structure(self, spreadsheet: str) -> dict:
        """生成可重复的结构摘要；按Sheet ID逐表读取A1:P10。"""
        sheets = self.query_sheets(spreadsheet)
        # Do not depend on batch range prefixes being IDs rather than titles.
        samples = {s['sheet_id']: self.read_values(spreadsheet, s['sheet_id'], 'A1:P10')
                   for s in sheets if s.get('sheet_id')}
        result = []
        for sheet in sheets:
            sheet_id = sheet.get('sheet_id') or ''
            if not sheet_id:
                continue
            grid = sheet.get('grid_properties') or {}
            result.append({
                'title': sheet.get('title') or '',
                'index': sheet.get('index'),
                'row_count': grid.get('row_count'),
                'column_count': grid.get('column_count'),
                'sample': samples.get(sheet_id, []),
            })
        canonical = json.dumps(result, ensure_ascii=False, sort_keys=True,
                               separators=(',', ':'), default=str)
        return {
            'sheet_count': len(result),
            'sheets': result,
            'sha256': hashlib.sha256(canonical.encode('utf-8')).hexdigest(),
        }

    def read_values_batch(self, spreadsheet: str, ranges: list[str]) -> dict[str, list]:
        """分批读取多个 range，返回 {sheet_id: values}。"""
        result = {}
        for start in range(0, len(ranges), 10):
            chunk = ranges[start:start + 10]
            params = [('ranges', item) for item in chunk]
            r = self._client.get(
                f'/sheets/v2/spreadsheets/{spreadsheet}/values_batch_get',
                params=params, headers=self._headers())
            r.raise_for_status()
            d = r.json()
            if d.get('code') != 0:
                raise RuntimeError(f"批量读取失败: {d.get('msg')}")
            for item in ((d.get('data') or {}).get('valueRanges') or []):
                value_range = item.get('valueRange') or item
                full_range = str(value_range.get('range') or '')
                sheet_id = full_range.split('!', 1)[0]
                values = value_range.get('values') or []
                if sheet_id:
                    result[sheet_id] = values
        return result

    def copy_file(self, file_token: str, file_type: str, name: str,
                  folder_token: str = '') -> dict:
        """复制云文档到指定 Drive 目录；空 folder_token 表示应用根目录。"""
        body = {'name': name, 'type': file_type, 'folder_token': folder_token}
        r = self._client.post(f'/drive/v1/files/{file_token}/copy',
                              json=body, headers=self._headers())
        r.raise_for_status()
        d = r.json()
        if d.get('code') != 0:
            raise RuntimeError(f"创建飞书副本失败: {d.get('msg')} (code={d.get('code')})")
        file_info = (d.get('data') or {}).get('file') or {}
        token = file_info.get('token') or ''
        if not token:
            raise RuntimeError('创建飞书副本成功但响应缺少副本 Token')
        return file_info

    def list_root_files(self, page_size: int = 50) -> list[dict]:
        """只读列出应用 Drive 根目录最近文件，用于找回已创建的 TEST PoC。"""
        r = self._client.get('/drive/v1/files', params={
            'page_size': page_size, 'order_by': 'CreatedTime', 'direction': 'DESC',
        }, headers=self._headers())
        r.raise_for_status()
        d = r.json()
        if d.get('code') != 0:
            raise RuntimeError(f"读取 Drive 根目录失败: {d.get('msg')}")
        return ((d.get('data') or {}).get('files') or [])

    def wait_spreadsheet_structure(self, spreadsheet: str,
                                   attempts: int = 5) -> dict:
        """等待刚复制的 Spreadsheet 可读；只重试暂时性服务端/传输错误。"""
        last_error = None
        for attempt in range(attempts):
            try:
                return self.spreadsheet_structure(spreadsheet)
            except (RuntimeError, httpx.TransportError, httpx.HTTPStatusError) as exc:
                last_error = exc
                if attempt + 1 < attempts:
                    time.sleep(2 + attempt * 2)
        raise RuntimeError(f'副本在有限轮询后仍不可读: {last_error}')

    def create_spreadsheet(self, title: str, folder_token: str = '') -> dict:
        """在 Drive 中创建独立 Spreadsheet。"""
        r = self._client.post('/sheets/v3/spreadsheets', json={
            'title': title, 'folder_token': folder_token,
        }, headers=self._headers())
        r.raise_for_status()
        d = r.json()
        if d.get('code') != 0:
            raise RuntimeError(f"创建 Spreadsheet 失败: {d.get('msg')} (code={d.get('code')})")
        spreadsheet = (d.get('data') or {}).get('spreadsheet') or {}
        token = spreadsheet.get('spreadsheet_token') or ''
        if not token:
            raise RuntimeError('创建 Spreadsheet 成功但响应缺少 spreadsheet_token')
        return spreadsheet

    def add_sheet(self, spreadsheet: str, title: str, index: int = 1) -> str:
        """在独立结果 Spreadsheet 新建一个结果子表并返回 Sheet ID。"""
        body = {'requests': [{'addSheet': {'properties': {
            'title': title, 'index': index,
        }}}]}
        r = self._client.post(
            f'/sheets/v2/spreadsheets/{spreadsheet}/sheets_batch_update',
            json=body, headers=self._headers())
        r.raise_for_status()
        d = r.json()
        if d.get('code') != 0:
            raise RuntimeError(f"创建结果子表失败: {d.get('msg')} (code={d.get('code')})")
        replies = (d.get('data') or {}).get('replies') or []
        props = ((replies[0].get('addSheet') or {}).get('properties')
                 if replies else {}) or {}
        sheet_id = props.get('sheetId') or props.get('sheet_id') or ''
        if not sheet_id:
            # 部分响应不返回 properties；从元数据按唯一标题回查。
            matches = [s.get('sheet_id') for s in self.query_sheets(spreadsheet)
                       if s.get('title') == title and s.get('sheet_id')]
            if len(matches) != 1:
                raise RuntimeError(f'创建后无法唯一定位结果子表: {title}')
            sheet_id = matches[0]
        return sheet_id

    def inspect_weekly_registry(self, registry_url: str,
                                configured_sheet_id: str = '') -> dict:
        """只读解析固定登记表并返回唯一登记 Sheet 的记录。"""
        spreadsheet, obj_type = self.resolve_wiki_obj(registry_url)
        if obj_type != 'sheet':
            raise RuntimeError(f'固定登记表必须指向电子表格，当前类型: {obj_type}')
        sheet_map = self.list_sheets(spreadsheet)
        if configured_sheet_id:
            candidates = [(title, sid) for title, sid in sheet_map.items()
                          if sid == configured_sheet_id]
            if not candidates:
                raise RuntimeError(f'配置的登记 Sheet ID 不存在: {configured_sheet_id}')
        else:
            candidates = list(sheet_map.items())

        matches = []
        errors = {}
        for title, sheet_id in candidates:
            try:
                values = self.read_values(spreadsheet, sheet_id, 'A1:Z500')
                records = parse_registry_values(values)
                matches.append((title, sheet_id, records))
            except RuntimeError as exc:
                errors[title] = str(exc)
        if len(matches) != 1:
            raise RuntimeError(
                f'必须识别到唯一登记子表，当前匹配 {len(matches)} 个；'
                f'扫描结果: {errors}'
            )
        title, sheet_id, records = matches[0]
        return {
            'spreadsheet_token': spreadsheet,
            'obj_type': obj_type,
            'sheet_title': title,
            'sheet_id': sheet_id,
            'records': records,
            'sheet_count': len(sheet_map),
        }

    def read_values(self, spreadsheet: str, sheet_id: str, rng: str) -> list[list]:
        r = self._client.get(
            f'/sheets/v2/spreadsheets/{spreadsheet}/values/{sheet_id}!{rng}',
            headers=self._headers())
        r.raise_for_status()
        d = r.json()
        if d.get('code') != 0:
            raise RuntimeError(f"读取失败 {sheet_id}!{rng}: {d.get('msg')}")
        return ((d.get('data') or {}).get('valueRange') or {}).get('values') or []

    def write_values(self, spreadsheet: str, sheet_id: str, rng: str,
                     values: list[list]) -> None:
        body = {'valueRange': {'range': f'{sheet_id}!{rng}', 'values': _json_safe(values)}}
        r = self._client.put(
            f'/sheets/v2/spreadsheets/{spreadsheet}/values',
            params={'valueInputOption': 'USER_ENTERED'},
            json=body, headers=self._headers())
        r.raise_for_status()
        d = r.json()
        if d.get('code') != 0:
            raise RuntimeError(f"写入失败 {sheet_id}!{rng}: {d.get('msg')}")

    def _target_header_row(self, spreadsheet: str, sheet_id: str, cfg: dict) -> int:
        """探测目标表含 ASIN 的表头行（1-based），找不到用 feishu_header_row 兜底"""
        vals = self.read_values(spreadsheet, sheet_id, 'A1:O6')
        hr0 = _detect_header_row(vals)
        return (hr0 + 1) if hr0 >= 0 else cfg['feishu_header_row']

    def _target_cols(self, spreadsheet: str, sheet_id: str, cfg: dict) -> dict:
        """目标表列位置：按表头名定位，找不到用 source_cols 兜底"""
        vals = self.read_values(spreadsheet, sheet_id, 'A1:O6')
        hr0 = _detect_header_row(vals)
        if hr0 >= 0:
            return _resolve_cols(vals[hr0], cfg)
        return dict(cfg['source_cols'])

    def resolve_output_layout(self, spreadsheet: str, sheets: list[str],
                              cfg: dict) -> dict[str, int]:
        """P0(2.2)：确定每个 Sheet 的六列起始列。
        优先级：已有六列表头（删列/移动后自动适配）→ 迁移记录 → 最后业务列后连续空列。
        返回 {sheet: start_col(1-based)}，并写入迁移记录。"""
        headers = cfg['feishu_output_headers']
        record = load_migration_layout()
        result: dict[str, int] = {}
        sheet_map = self.list_sheets(spreadsheet)
        for sheet in sheets:
            sid = sheet_map.get(sheet)
            if not sid:
                continue
            hr = self._target_header_row(spreadsheet, sid, cfg)
            vals = self.read_values(spreadsheet, sid, f'A1:ZZ{min(hr + 3, 20)}')
            header_row = vals[hr - 1] if len(vals) >= hr else []
            # 1. 已有六列表头优先（表头在删列/移动后位置变化也能适配）
            existing = _find_existing_six(header_row, headers)
            if existing:
                result[sheet] = existing
                continue
            # 2. 迁移记录
            if sheet in record:
                result[sheet] = record[sheet]
                continue
            # 3. 自动探测空列
            start = resolve_sheet_start_col(header_row, vals, headers)
            if start:
                result[sheet] = start
        if result:
            save_migration_layout(result)
        return result

    def build_asin_map(self, spreadsheet: str, sheet_id: str) -> dict[str, int]:
        """读 A 列定位 ASIN 行（数据从表头行+1 开始，表头自动探测）"""
        from sheet_io import read_rows
        hr = self._target_header_row(spreadsheet, sheet_id, self.cfg)
        info = next((item for item in self.query_sheets(spreadsheet)
                     if item.get('sheet_id') == sheet_id), {})
        vals = read_rows(self, spreadsheet, sheet_id, last='A', start=hr + 1,
                         row_count=(info.get('grid_properties') or {}).get('row_count'))
        m: dict[str, int] = {}
        for i, cell in enumerate(vals):
            v = (cell[0] if isinstance(cell, list) and cell else cell) or ''
            s = str(v).strip()
            if s.startswith('B0'):
                m[s] = hr + 1 + i
        return m

    # ---------- 原始周报读取 ----------
    def read_source_sheets(self, spreadsheet: str, sheets: list[str],
                           cfg: dict) -> tuple[dict[str, list[ReportRow]], dict, list[dict]]:
        """读原始表各子表 → (ReportRow 列表, source_meta, 无效行列表)。
        表头自动探测 + 列名定位（兼容不同 Sheet 布局，如 PD05 目标价在 L 列）。
        3.6 I 列按百分比语义解析（'20%' → 0.20）。
        3.7 目标成交价/正常售价为空 → 无效行，不进入正式抓取。"""
        from sheet_io import read_rows
        out: dict[str, list[ReportRow]] = {}
        invalid_all: list[dict] = []
        meta = {'spreadsheet': spreadsheet, 'read_at': time.strftime('%Y-%m-%d %H:%M:%S'),
                'sheets': {}}
        sheet_map = self.list_sheets(spreadsheet)
        metadata = {item.get('sheet_id'): item for item in self.query_sheets(spreadsheet)}
        for sheet in sheets:
            sid = sheet_map.get(sheet)
            if not sid:
                continue
            info = metadata.get(sid) or {}
            grid = info.get('grid_properties') or {}
            vals = read_rows(self, spreadsheet, sid,
                             last=col_letter(max(15, int(grid.get('column_count') or 15))),
                             row_count=grid.get('row_count'))
            try:
                rows, invalid = read_source_rows(vals, cfg)
            except RuntimeError as e:
                raise RuntimeError(f'[{sheet}] {e}')
            for iv in invalid:
                iv['sheet'] = sheet
            invalid_all.extend(invalid)
            if rows:
                out[sheet] = rows
                meta['sheets'][sheet] = {'rows': len(rows), 'invalid': len(invalid)}
        return out, meta, invalid_all

    def backup_target_sheet(self, spreadsheet: str, sheet: str, sheet_id: str,
                            run_id: str) -> Path:
        """覆盖紧凑目标表前保存 A:V 快照，供迁移/误操作恢复。"""
        from config import OUTPUT_DIR
        import json
        path = OUTPUT_DIR / 'target_backups' / run_id / f'{sheet}.json'
        if path.is_file():
            saved = json.loads(path.read_text(encoding='utf-8'))
            if saved.get('sheet') != sheet or not isinstance(saved.get('values'), list):
                raise RuntimeError(f'原始备份无效，停止覆盖: {path}')
            if saved.get('spreadsheet_token', spreadsheet) != spreadsheet:
                raise RuntimeError('备份目标与当前结果表不一致')
            return path  # Preserve the pre-write image across retries of the same run.
        from sheet_io import read_rows
        from runtime_state import atomic_json
        info = next((s for s in self.query_sheets(spreadsheet) if s['sheet_id'] == sheet_id), {})
        capacity = (info.get('grid_properties') or {}).get('row_count') or 2000
        values = read_rows(self, spreadsheet, sheet_id, last='V', row_count=capacity)
        path.parent.mkdir(parents=True, exist_ok=True)
        atomic_json(path, {'sheet': sheet, 'sheet_id': sheet_id, 'spreadsheet_token': spreadsheet,
                          'range': f'A1:V{capacity}', 'values': values})
        return path

    # ---------- 紧凑基础数据同步：A:G 原始字段，H:M 抓取结果 ----------
    def sync_base_data(self, spreadsheet: str, sheet: str, sheet_id: str,
                       rows: list[ReportRow], cfg: dict) -> tuple[int, int, int]:
        """每天完整刷新紧凑目标表。

        第2行固定表头；A:G 是源数据，H:M 是抓取六列。同步前清空旧数据，
        避免源表换行后旧抓取结果挂到错误 ASIN。
        返回 (写入单元格数, 当前行数, 被移除旧行数)。
        """
        hr = 2
        data_start = 3
        output_headers = list(cfg['feishu_output_headers'])
        headers = COMPACT_BASE_HEADERS + output_headers

        old_map = self.build_asin_map(spreadsheet, sheet_id)
        old_count = len(old_map)
        clear_end = max(data_start + len(rows) - 1,
                        max(old_map.values()) if old_map else data_start)

        # 找到并清除旧六列区域（例如 PD03 的 Q:V）。只有表头完整匹配时才清理。
        old_header_vals = self.read_values(spreadsheet, sheet_id, f'A{hr}:ZZ{hr}')
        old_header = old_header_vals[0] if old_header_vals else []
        old_output_starts = []
        for idx in range(0, max(0, len(old_header) - len(output_headers) + 1)):
            if all(isinstance(old_header[idx + j], str)
                   and old_header[idx + j].strip() == output_headers[j]
                   for j in range(len(output_headers))):
                old_output_starts.append(idx + 1)
        for old_output_start in (c for c in old_output_starts if c != 8):
            old_end_col = old_output_start + 5
            for start in range(hr, clear_end + 1, 200):
                end = min(start + 199, clear_end)
                self.write_values(
                    spreadsheet, sheet_id,
                    f'{col_letter(old_output_start)}{start}:{col_letter(old_end_col)}{end}',
                    [[''] * 6 for _ in range(end - start + 1)],
                )

        # 固定统一表头 A:M。
        self.write_values(spreadsheet, sheet_id, f'A{hr}:M{hr}', [headers])

        # 清除旧 A:M 数据，分批写空值，避免 API 单次范围过大。
        for start in range(data_start, clear_end + 1, 200):
            end = min(start + 199, clear_end)
            self.write_values(spreadsheet, sheet_id, f'A{start}:M{end}',
                              [[''] * 13 for _ in range(end - start + 1)])

        def _i_display(v):
            if isinstance(v, Decimal) and 0 < v < 1:
                text = format(v * 100, 'f').rstrip('0').rstrip('.')
                return f'{text}%'
            return v if v is not None else ''

        values = [[
            r.asin,
            r.sku,
            r.size,
            r.normal_price,
            r.h_type,
            _i_display(r.i_value),
            r.target_price,
        ] for r in rows]
        if values:
            end = data_start + len(values) - 1
            self.write_values(spreadsheet, sheet_id, f'A{data_start}:G{end}', values)

        removed = max(0, old_count - len(rows))
        written = 13 + len(values) * 7
        return written, len(rows), removed

    # ---------- 六列写入（稀疏、只写有结果行） ----------
    def write_six_columns(self, spreadsheet: str, sheet: str, sheet_id: str,
                          asin_map: dict[str, int], crawls: list[CrawlResult],
                          cfg: dict, start_col: int | None = None) -> int:
        start = start_col or cfg['feishu_output_start_col']
        end_col = start + 5
        rows_out: dict[int, list] = {}
        for cr in crawls:
            row = asin_map.get(cr.asin)
            if row is None:
                continue
            rows_out[row] = cr.six_columns()
        if not rows_out:
            return 0
        rng_letters = f'{col_letter(start)}:{col_letter(end_col)}'
        written = 0
        for start_row, end_row, values in _group_contiguous_rows(rows_out):
            self.write_values(spreadsheet, sheet_id,
                              f'{col_letter(start)}{start_row}:{col_letter(end_col)}{end_row}',
                              values)
            written += len(values)
        return written

    # ---------- 布局只读预检（3.9 / P0-2.3） ----------
    def inspect_layout(self, spreadsheet: str, sheets: list[str], cfg: dict) -> dict:
        """只读检查目标表布局：表头行/ASIN起始行/目标价列/最后业务列/旧列/建议输出范围/覆盖风险。不写数据。"""
        headers = cfg['feishu_output_headers']
        result: dict[str, dict] = {}
        sheet_map = self.list_sheets(spreadsheet)
        for sheet in sheets:
            sid = sheet_map.get(sheet)
            if not sid:
                result[sheet] = {'error': '目标表无此 sheet'}
                continue
            hr = self._target_header_row(spreadsheet, sid, cfg)
            top = self.read_values(spreadsheet, sid, f'A1:O{min(6, hr + 2)}')
            asin_rows = self.read_values(spreadsheet, sid, f'A{hr + 1}:A300')
            first_asin = None
            for i, cell in enumerate(asin_rows):
                v = (cell[0] if isinstance(cell, list) and cell else cell) or ''
                if str(v).strip().startswith('B0'):
                    first_asin = {'row': hr + 1 + i, 'asin': str(v).strip()}
                    break
            header_row = (top[hr - 1] if len(top) >= hr else []) or []
            legacy_cols = [i + 1 for i, v in enumerate(header_row)
                           if isinstance(v, str) and v.strip() in LEGACY_HEADERS]
            # 目标成交价列 / 最后业务列
            tgt_col = None
            last_biz = 0
            for i, v in enumerate(header_row):
                if isinstance(v, str) and v.strip():
                    last_biz = i + 1
                    if v.strip() == '目标成交价':
                        tgt_col = i + 1
            # 建议输出范围
            full = self.read_values(spreadsheet, sid, f'A1:ZZ{min(hr + 3, 20)}')
            suggested = resolve_sheet_start_col(header_row, full, headers)
            # 已存在六列？
            existing = None
            for i, cell in enumerate(header_row):
                if isinstance(cell, str) and cell.strip() == headers[0]:
                    if all(isinstance(header_row[i + j], str)
                           and header_row[i + j].strip() == headers[j]
                           for j in range(len(headers)) if i + j < len(header_row)):
                        existing = i + 1
                        break
            # J:O 当前内容（固定 J 列检查，用于人工判断）
            jo = self.read_values(spreadsheet, sid, 'J1:O20')
            jo_nonempty = []
            for ri, row in enumerate(jo):
                for ci, v in enumerate(row):
                    if v is not None and str(v).strip() != '':
                        jo_nonempty.append({'cell': f'{col_letter(10 + ci)}{1 + ri}',
                                            'value': str(v)[:20]})
            overlap_risk = ('有' if (tgt_col and existing and existing <= tgt_col + 1)
                            else ('无' if (suggested and suggested > last_biz) else '需人工确认'))
            result[sheet] = {
                'header_row': hr,
                'first_asin': first_asin,
                'target_price_col': f'{col_letter(tgt_col)}列' if tgt_col else '未发现',
                'last_business_col': f'{col_letter(last_biz)}列' if last_biz else '无',
                'legacy_cols': legacy_cols,
                'existing_six_cols': f'{col_letter(existing)}列' if existing else '未发现',
                'suggested_output': (f'{col_letter(suggested)}:{col_letter(suggested + 5)}'
                                     if suggested else '找不到可用区域'),
                'j_o_nonempty': jo_nonempty[:20],
                'overlap_risk': overlap_risk,
            }
        return result

    # ---------- 一次性列迁移 ----------
    def migrate_columns(self, spreadsheet: str, cfg: dict, sheets: list[str],
                        confirm: bool = False, force: bool = False) -> dict:
        """清理旧追加列数据 → 在解析出的起始列写六列表头。要求显式 confirm。
        安全（P0-2.4）：输出区域非旧字段拒绝迁移；迁移记录防止重复执行。"""
        if not confirm:
            raise RuntimeError('列迁移是破坏性一次性操作，必须加 --confirm 才能执行')
        from config import OUTPUT_DIR
        record_path = OUTPUT_DIR / 'migration_record.json'
        if record_path.exists() and not force:
            raise RuntimeError(
                f'检测到已执行过列迁移（{record_path}）。如确需重跑请加 --force 覆盖记录。')
        headers = cfg['feishu_output_headers']
        result: dict[str, dict] = {}
        sheet_map = self.list_sheets(spreadsheet)
        layout: dict[str, int] = {}

        # 1. 迁移前快照（表头 + 输出区数据）
        import json
        backup: dict[str, dict] = {'captured_at': time.strftime('%Y-%m-%d %H:%M:%S')}
        (OUTPUT_DIR / 'migration_backup').mkdir(parents=True, exist_ok=True)

        for sheet in sheets:
            sid = sheet_map.get(sheet)
            if not sid:
                print(f'  [{sheet}] 目标表无此 sheet，跳过', flush=True)
                continue
            hr = self._target_header_row(spreadsheet, sid, cfg)
            # 0. 解析输出起始列（已有六列复用 / 最后业务列后空列）
            vals = self.read_values(spreadsheet, sid, f'A1:ZZ{min(hr + 3, 20)}')
            header_row = vals[hr - 1] if len(vals) >= hr else []
            start = resolve_sheet_start_col(header_row, vals, headers)
            if start is None:
                raise RuntimeError(f'[{sheet}] 找不到可用的六列输出区域（业务列已占满表宽）')
            layout[sheet] = start

            # 1. 读表头行，定位旧追加列
            legacy_cols = [i + 1 for i, v in enumerate(header_row)
                           if isinstance(v, str) and v.strip() in LEGACY_HEADERS]

            # 4. 检查输出区域是否存在非旧字段数据，存在则拒绝迁移
            jo_vals = self.read_values(spreadsheet, sid, f'{col_letter(start)}{hr}:{col_letter(start + 5)}30')
            non_legacy = []
            for ri, row in enumerate(jo_vals):
                for ci, v in enumerate(row):
                    if v is None or str(v).strip() == '':
                        continue
                    col = start + ci
                    if ri == 0:                      # 表头行：已存在六列新表头
                        if str(v).strip() in headers:
                            continue
                    if col in legacy_cols:
                        continue
                    non_legacy.append(f'{col_letter(col)}{hr + ri}')
            if non_legacy:
                raise RuntimeError(
                    f'[{sheet}] 输出区域 {col_letter(start)}:{col_letter(start + 5)} 存在非旧字段数据 '
                    f'{non_legacy[:10]}，拒绝迁移。请先人工确认这些数据。')

            backup[sheet] = {'header': header_row[:20],
                             'j_o': [[str(v)[:20] if v is not None else '' for v in row]
                                     for row in jo_vals[:10]]}
            # 2. 清除旧列数据（保留表头；range 行数与值数量必须精确匹配）
            for lc in legacy_cols:
                clear_rows = 400                       # 覆盖数据区 + 余量
                self.write_values(spreadsheet, sid,
                                  f'{col_letter(lc)}{hr + 1}:{col_letter(lc)}{hr + clear_rows}',
                                  [['']] * clear_rows)
            # 3. 写六列表头
            self.write_values(spreadsheet, sid,
                              f'{col_letter(start)}{hr}:{col_letter(start + 5)}{hr}',
                              [headers])
            result[sheet] = {'legacy_cleared': legacy_cols,
                             'start_col': start,
                             'output_range': f'{col_letter(start)}{hr}:{col_letter(start + 5)}'}
            print(f'  [{sheet}] 已迁移: 清除旧列 {legacy_cols}, 六列表头写入 '
                  f'{col_letter(start)}{hr}:{col_letter(start + 5)}{hr}', flush=True)

        # 5/6/7. 本地迁移记录 + backup
        import json as _json
        backup_path = OUTPUT_DIR / 'migration_backup' / f'before_{time.strftime("%Y%m%d_%H%M%S")}.json'
        with open(backup_path, 'w', encoding='utf-8') as f:
            _json.dump(backup, f, ensure_ascii=False, indent=1)
        with open(record_path, 'w', encoding='utf-8') as f:
            _json.dump({
                'completed_at': time.strftime('%Y-%m-%d %H:%M:%S'),
                'sheets': list(result.keys()),
                'start_col': start,                    # 最后一个 sheet 的起始列（参考）
                'headers': headers,
                'backup': str(backup_path),
                'output_layout': layout,               # P0-2.2：每 Sheet 起始列，日常运行优先读
            }, f, ensure_ascii=False, indent=1)
        save_migration_layout(layout)
        print(f'[迁移] 备份与记录: {backup_path} / {record_path}', flush=True)
        print(f'[迁移] 每 Sheet 起始列: {layout}', flush=True)
        return result

    def close(self):
        try:
            self._client.close()
        except Exception:
            pass


def _parse_i_value(v):
    """3.6 I 列（本周折扣值）按列语义解析：'20%' → Decimal('0.20')；普通数值按原值。"""
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        if s.endswith('%'):
            try:
                return Decimal(s[:-1]) / Decimal(100)
            except (InvalidOperation, ZeroDivisionError):
                return None
    return _num_or_none(v)


def _group_contiguous(pairs: list[tuple[int, object]]):
    """[(row, value)] → 连续段 [(start, end, [values])]"""
    pairs = sorted(pairs, key=lambda x: x[0])
    if not pairs:
        return []
    segs = []
    cur_start, cur_end = pairs[0][0], pairs[0][0]
    cur_vals = [pairs[0][1]]
    for row, v in pairs[1:]:
        if row == cur_end + 1:
            cur_end = row
            cur_vals.append(v)
        else:
            segs.append((cur_start, cur_end, cur_vals))
            cur_start = cur_end = row
            cur_vals = [v]
    segs.append((cur_start, cur_end, cur_vals))
    return segs


def _group_contiguous_rows(rows_out: dict[int, list]):
    """{row: values} → 连续段 [(start, end, [values...])]，每段一个 range"""
    items = sorted(rows_out.items())
    if not items:
        return []
    segs = []
    cur_start = items[0][0]
    cur_vals = [items[0][1]]
    prev = items[0][0]
    for row, v in items[1:]:
        if row == prev + 1:
            cur_vals.append(v)
        else:
            segs.append((cur_start, prev, cur_vals))
            cur_start = row
            cur_vals = [v]
        prev = row
    segs.append((cur_start, prev, cur_vals))
    return segs
